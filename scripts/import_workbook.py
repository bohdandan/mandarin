from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.vocabulary import ImportResult, build_entry, strip_html, write_source_vocabulary, write_vocabulary


def row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {header: row[index] if index < len(row) else "" for index, header in enumerate(headers)}


def looks_sentence_like(value: str) -> bool:
    plain = strip_html(value)
    return bool("<" in value or ">" in value or len(repr(plain)) > 28 or any(char in plain for char in "。！？.!?"))


def import_workbook(workbook_path: Path) -> ImportResult:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    entries: list[dict[str, Any]] = []
    skipped_duplicates: list[str] = []
    suspicious_rows: list[str] = []
    seen_hanzi: set[str] = set()

    if "HSK 1-4" in wb.sheetnames:
        ws = wb["HSK 1-4"]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        for row_number, row in enumerate(rows, start=2):
            data = row_dict(headers, row)
            if not any(data.values()):
                continue
            entry = build_entry(
                raw_id=data.get("id"),
                hanzi=data.get("hanzi"),
                pinyin=data.get("pinyin"),
                english=data.get("english"),
                example_sentence=data.get("example_sentence"),
                sentence_pinyin=data.get("sentence_pinyin"),
                source="hsk-workbook",
                tags_raw=str(data.get("tags") or ""),
            )
            if not entry["hanzi"]:
                suspicious_rows.append(f"HSK 1-4 row {row_number}: missing hanzi")
                continue
            entries.append(entry)
            seen_hanzi.add(entry["hanzi"])

    if "Custom" in wb.sheetnames:
        ws = wb["Custom"]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        for row_number, row in enumerate(rows, start=2):
            data = row_dict(headers, row)
            hanzi = str(data.get("hanzi") or "").strip()
            if not hanzi:
                continue
            translation = str(data.get("translation") or "").strip()
            if hanzi in seen_hanzi:
                skipped_duplicates.append(hanzi)
                continue
            if looks_sentence_like(translation):
                suspicious_rows.append(f"Custom row {row_number} ({hanzi}): sentence-like translation")
            entry = build_entry(
                hanzi=hanzi,
                pinyin=data.get("pinyin"),
                english=translation,
                example_sentence=data.get("example_sentence"),
                source="custom",
                lesson="custom",
                tags_raw="",
            )
            entries.append(entry)
            seen_hanzi.add(hanzi)

    return ImportResult(entries=entries, skipped_duplicates=skipped_duplicates, suspicious_rows=suspicious_rows)


def write_import_log(result: ImportResult, log_path: Path) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Workbook Import",
        "",
        f"Imported entries: {len(result.entries)}",
        f"Skipped duplicates: {len(result.skipped_duplicates)}",
        f"Suspicious rows: {len(result.suspicious_rows)}",
        "",
    ]
    if result.skipped_duplicates:
        lines.extend(["## Skipped Duplicates", "", *[f"- {word}" for word in result.skipped_duplicates], ""])
    if result.suspicious_rows:
        lines.extend(["## Suspicious Rows", "", *[f"- {row}" for row in result.suspicious_rows], ""])
    log_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import the HSK workbook into canonical vocabulary JSON.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--output-dir", type=Path, default=Path("data/sources"))
    parser.add_argument("--log", type=Path, default=Path("logs/imports/latest.md"))
    args = parser.parse_args()

    result = import_workbook(args.workbook)
    write_source_vocabulary(args.output_dir, result.entries)
    if args.output is not None:
        write_vocabulary(args.output, result.entries)
    write_import_log(result, args.log)

    print(f"Imported {len(result.entries)} entries into {args.output_dir}")
    print(f"Skipped {len(result.skipped_duplicates)} duplicate custom rows")
    print(f"Flagged {len(result.suspicious_rows)} suspicious rows")
    if result.suspicious_rows:
        for row in result.suspicious_rows:
            print(f"WARNING: {row}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
