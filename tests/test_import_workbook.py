import tempfile
import unittest
import subprocess
import sys
from pathlib import Path

import openpyxl

from scripts.import_workbook import import_workbook


class ImportWorkbookTest(unittest.TestCase):
    def build_sample_workbook(self, workbook_path: Path) -> None:
        wb = openpyxl.Workbook()
        main = wb.active
        main.title = "HSK 1-4"
        main.append(["id", "hanzi", "pinyin", "english", "example_sentence", "sentence_pinyin", "tags"])
        main.append([1, "你好", "nǐ hǎo", "hello", "<b>你好</b>！", "nǐ hǎo", "HSK1::HSK:1.01"])
        custom = wb.create_sheet("Custom")
        custom.append(["hanzi", "pinyin", "translation", "example_sentence", ""])
        custom.append(["杯子", "bēi zi", "cup", "这是我的<b>杯子</b>。", ""])
        wb.save(workbook_path)

    def test_imports_main_rows_and_skips_duplicate_custom_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "sample.xlsx"
            wb = openpyxl.Workbook()
            main = wb.active
            main.title = "HSK 1-4"
            main.append(["id", "hanzi", "pinyin", "english", "example_sentence", "sentence_pinyin", "tags"])
            main.append([1, "你好", "nǐ hǎo", "hello", "<b>你好</b>！", "nǐ hǎo", "HSK1::HSK:1.01"])
            main.append([2, "手表", "shǒu biǎo", "watch", "这是我的<b>手表</b>。", "zhè shì wǒ de shǒu biǎo", "HSK2::HSK:2.04"])
            custom = wb.create_sheet("Custom")
            custom.append(["hanzi", "pinyin", "translation", "example_sentence", ""])
            custom.append(["你好", "nǐ hǎo", "hello again", "<b>你好</b>！", ""])
            custom.append(["杯子", "bēi zi", "cup", "这是我的<b>杯子</b>。", ""])
            wb.save(workbook_path)

            result = import_workbook(workbook_path)

        self.assertEqual([entry["hanzi"] for entry in result.entries], ["你好", "手表", "杯子"])
        self.assertEqual(result.skipped_duplicates, ["你好"])
        self.assertEqual(result.entries[2]["source"], "custom")
        self.assertEqual(result.entries[2]["hsk_level"], None)

    def test_flags_custom_rows_with_sentence_like_translation(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "sample.xlsx"
            wb = openpyxl.Workbook()
            main = wb.active
            main.title = "HSK 1-4"
            main.append(["id", "hanzi", "pinyin", "english", "example_sentence", "sentence_pinyin", "tags"])
            custom = wb.create_sheet("Custom")
            custom.append(["hanzi", "pinyin", "translation", "example_sentence", ""])
            custom.append(["地", "de", "他高兴<b>地</b>笑了。", "", ""])
            wb.save(workbook_path)

            result = import_workbook(workbook_path)

        self.assertEqual(len(result.suspicious_rows), 1)
        self.assertIn("sentence-like translation", result.suspicious_rows[0])

    def test_import_script_runs_directly_from_repo_root(self):
        repo_root = Path(__file__).resolve().parents[1]
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "sample.xlsx"
            output_dir = Path(tmp) / "sources"
            log_path = Path(tmp) / "import.md"
            self.build_sample_workbook(workbook_path)

            result = subprocess.run(
                [
                    sys.executable,
                    "scripts/import_workbook.py",
                    str(workbook_path),
                    "--output-dir",
                    str(output_dir),
                    "--log",
                    str(log_path),
                ],
                cwd=repo_root,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue((output_dir / "hsk-workbook.json").exists())
            self.assertTrue((output_dir / "custom.json").exists())


if __name__ == "__main__":
    unittest.main()
